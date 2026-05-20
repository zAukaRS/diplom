import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUTPUT_DIR = 'excel_files'

def _border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _header_style(cell):
    cell.font = Font(bold=True, name="Times New Roman", size=11)
    cell.fill = PatternFill("solid", fgColor="CCCCCC")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _border()

def _cell_style(cell, fill_color=None):
    cell.font = Font(name="Times New Roman", size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _border()
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

def _ensure_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def create_report(dict_list: list, output_filename: str, sheet_name='расч.л'):
    """
    Основной отчёт: кто жил, сколько дней, расчёт.

    dict_list: [
        {
            'Месторождение': str,
            'Заказчик': str,
            'ФИО проживающего': str,
            'Дата заезда': date,
            'Дата выезда': date,
            'Количество дней': int,
        }, ...
    ]
    """
    _ensure_dir()
    filepath = os.path.join(OUTPUT_DIR, output_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.row_dimensions[1].height = 35

    headers = [
        'Месторождение', 'Заказчик', 'ФИО проживающего',
        'Дата заезда', 'Дата выезда', 'Количество дней', 'Расчёт за жильё'
    ]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)

    for row in dict_list:
        days = int(row['Количество дней'])
        ws.append([
            row['Месторождение'],
            row['Заказчик'],
            row['ФИО проживающего'],
            row['Дата заезда'].strftime("%d.%m.%Y") if isinstance(row['Дата заезда'], date) else row['Дата заезда'],
            row['Дата выезда'].strftime("%d.%m.%Y") if isinstance(row['Дата выезда'], date) else row['Дата выезда'],
            days,
            days * 500,
        ])

    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            color = "86D472" if i in (4, 5, 6) else ("FFFF00" if i == 7 else None)
            _cell_style(cell, color)

    for i, width in enumerate([25, 25, 30, 15, 15, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filepath)
    return filepath


def over_time(dict_list: list, output_filename: str, date_from: date, date_to: date, norm_days: int = 15):
    """
    Отчёт по переработке: кто превысил норму дней проживания.

    dict_list: [
        {
            'Месторождение': str,
            'Заказчик': str,
            'ФИО проживающего': str,
            'Количество дней': int,   ← суммарные дни за период из SQL
        }, ...
    ]
    norm_days: норма дней (по умолчанию 15), переработка = дни - норма
    """
    _ensure_dir()
    filepath = os.path.join(OUTPUT_DIR, output_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = 'переработка'
    ws.row_dimensions[1].height = 40

    headers = [
        'Месторождение',
        'Заказчик',
        'ФИО проживающего',
        f'с {date_from.strftime("%d.%m.%Y")}',
        f'по {date_to.strftime("%d.%m.%Y")}',
        'Дней отработано',
        f'Норма ({norm_days} дн.)',
        'Переработка (дн.)',
        'Переработка (%)',
    ]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)

    # Строка итогов в конце
    total_days = 0
    total_over = 0

    for row in dict_list:
        days = int(row['Количество дней'])
        overtime = max(0, days - norm_days)
        pct = round((overtime / norm_days) * 100, 1) if norm_days else 0
        total_days += days
        total_over += overtime

        ws.append([
            row['Месторождение'],
            row['Заказчик'],
            row['ФИО проживающего'],
            date_from.strftime("%d.%m.%Y"),
            date_to.strftime("%d.%m.%Y"),
            days,
            norm_days,
            overtime,
            f"{pct}%",
        ])

    # Строка итогов
    total_row_idx = ws.max_row + 1
    ws.cell(total_row_idx, 1, 'ИТОГО')
    ws.cell(total_row_idx, 6, total_days)
    ws.cell(total_row_idx, 8, total_over)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
        for i, cell in enumerate(row, start=1):
            if i == 6:
                color = "86D472"   # дней отработано — зелёный
            elif i == 7:
                color = "CCCCCC"   # норма — серый
            elif i == 8:
                # переработка: красный если > 0, иначе белый
                color = "FF9999" if (cell.value or 0) > 0 else "FFFFFF"
            elif i == 9:
                color = "FFFF00"   # % — жёлтый
            else:
                color = None
            _cell_style(cell, color)

    # Стиль строки итогов
    for cell in ws[total_row_idx]:
        cell.font = Font(bold=True, name="Times New Roman", size=11)
        cell.fill = PatternFill("solid", fgColor="CCCCCC")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border()

    for i, width in enumerate([25, 25, 30, 15, 15, 18, 15, 18, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filepath)
    return filepath