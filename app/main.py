from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pathlib import Path as PathLib
from starlette.templating import Jinja2Templates
from .utils import get_admin_role_id
from .core.security import get_password_hash
from .core.dependencies import get_current_user
from .routers import auth, users
from sqlalchemy.orm import selectinload 
from sqlalchemy import text, and_
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi.responses import JSONResponse
from fastapi import File, UploadFile, Depends
import pandas as pd
from .database import get_db, Session_async

from datetime import datetime,timedelta, date
from fastapi import Body
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
from app.models import (
    User, Role, Field, Customer, Location, Path, Workplace, Room, Resident, ResidentDay
)
from sqlalchemy.orm import joinedload
from sqlalchemy import text, and_, or_, select, func, desc, extract, and_
from fastapi import Depends, HTTPException
import io
from .excel_parser import parse_all_months




# uvicorn app.main:app --reload

app = FastAPI()
app.include_router(auth.router)
app.include_router(users.router)

BASE_DIR = PathLib(__file__).parent.parent
FRONTEND_DIR = BASE_DIR / "frontend"
templates = Jinja2Templates(directory=FRONTEND_DIR)

app.mount("/css", StaticFiles(directory=FRONTEND_DIR / "css"), name="css")
app.mount("/js", StaticFiles(directory=FRONTEND_DIR / "js"), name="js")


def create_report(dict_list: list, output_filename: str, sheet_name='расч.л'):

    if not os.path.exists('excel_files'):
        os.makedirs('excel_files')

    filepath = os.path.join('excel_files', output_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = ['Месторождение', 'Заказчик', 'ФИО проживающего',
               'Дата заезда', 'Дата выезда', 'Количество дней']

    # добавляем данные
    ws.append(headers)

    for row in dict_list:
        ws.append([
            row['Месторождение'],
            row['Заказчик'],
            row['ФИО проживающего'],
            row['Дата заезда'].strftime("%d.%m.%Y"),
            row['Дата выезда'].strftime("%d.%m.%Y") if row['Дата выезда'] else "—",
            row['Количество дней']
        ])

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Times New Roman", size=11)
        cell.fill = PatternFill("solid", fgColor="CCCCCC")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Данные
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            if i in (4, 5, 6):
                cell.fill = PatternFill(start_color="86D472", end_color="86D472", fill_type="solid")

    widths = [25, 25, 30, 15, 15, 20]

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filepath)
    return filepath



# get_current_user теперь из .core.dependencies (JWT Bearer)

async def admin_only(user: User = Depends(get_current_user)):
    if user.role.name != "admin":
        raise HTTPException(status_code=403, detail="Доступ запрещён")
    return user



def parse_date_dd_mm_yyyy(date_str):

    if pd.isna(date_str):
        return None
    if isinstance(date_str, datetime):
        return date_str.date()
    try:
        return datetime.strptime(str(date_str).strip(), "%d.%m.%Y").date()
    except ValueError:
        return None


def search(word : str, data : list):
    word = word.lower()
    filters = ['full_name','position','room_number','field','room_location']
    for i in filters:
        res = [{"id": r.id,
                "room_number": r.room_number or "",
                "room_location": r.room_location or "",
                "room_path": r.room_path or "",
                "full_name": r.full_name ,
                "position": r.position or "",
                "gender": r.gender or "",
                "shift": r.shift or "",
                "field": r.field ,
                "customer": r.customer,
                "days_info": r.days_info} for r in data if r[i] == f'{word}%']
        if res:
            return res


@app.get("/login", response_class=HTMLResponse)
async def login_page():
    return HTMLResponse((FRONTEND_DIR / "login.html").read_text(encoding="utf-8"))


# POST /login и /logout заменены роутером /api/auth/login и /api/auth/refresh



# Главная страница
@app.get("/home", response_class=HTMLResponse)
def home():
    return HTMLResponse((FRONTEND_DIR / "index.html").read_text(encoding="utf-8"))

# Чтобы заход на / сразу редиректил на /login
@app.get("/", include_in_schema=False)
def root():
    return RedirectResponse(url="/login")



@app.post("/api/update_day")
async def update_day(data: dict = Body(...), db: AsyncSession = Depends(get_db)):
    try:
        resident_id = int(data["resident_id"])
        day = int(data["day"])
        month = int(data["month"])
        year = int(data["year"])
        customer_id = data.get("customer_id")
        if customer_id is not None:
            customer_id = int(customer_id)

        target_date = date(year, month, day)

        # Ищем существующую запись
        res = await db.execute(
            select(ResidentDay).where(
                and_(
                    ResidentDay.resident_id == resident_id,
                    ResidentDay.date == target_date
                )
            )
        )
        rd = res.scalars().first()

        if rd:
            # Обновляем только заказчика
            rd.customer_id = customer_id
        else:
            # Нужно получить данные из родительского Resident
            res_r = await db.execute(
                select(Resident).where(Resident.id == resident_id)
            )
            resident = res_r.scalars().first()
            if not resident:
                raise HTTPException(404, f"Проживающий {resident_id} не найден")

            rd = ResidentDay(
                resident_id=resident_id,
                room_id=resident.room_id,     
                date=target_date,
                extra=resident.check_out,      
                customer_id=customer_id
            )
            db.add(rd)

        await db.commit()
        return {"status": "ok"}

    except HTTPException as he:
        raise he
    except Exception as e:
        await db.rollback()
        raise HTTPException(400, detail=str(e))
    




@app.get("/api/residents")
async def get_residents(
    word: Optional[str] = None, 
    by_field: Optional[str] = None, 
    db: AsyncSession = Depends(get_db)
):
    
    
    query = select(Resident).options(
        selectinload(Resident.field),
        selectinload(Resident.customer),
        selectinload(Resident.room).selectinload(Room.location),
        selectinload(Resident.room).selectinload(Room.path),
        selectinload(Resident.resident_days)
    )
    
    # Фильтр по ID месторождения
    if by_field and by_field.strip():
        try:
            field_id = int(by_field)
            query = query.where(Resident.field_id == field_id)
            
        except ValueError:
            pass
    
    # Фильтр по поисковому слову
    if word:
        word_lower = word.lower().strip()
        query = query.outerjoin(Room)\
                     .outerjoin(Location)\
                     .outerjoin(Customer)\
                     .where(
            or_(
                Resident.full_name.ilike(f"%{word_lower}%"),
                Resident.position.ilike(f"%{word_lower}%"),
                Room.room_number.ilike(f"%{word_lower}%"),
                Location.name.ilike(f"%{word_lower}%"),
                Customer.name.ilike(f"%{word_lower}%")
            )
        )

    
    # Выполняем запрос
    result = await db.execute(query)
    residents = result.scalars().unique().all()
    
   
    
    if not residents:
        return []
    
    # Формируем ответ
    response = []
    for r in residents:
        room = r.room
        location = room.location if room else None
        
        days_info = {}
        for rd in r.resident_days:
            days_info[rd.date.day] = rd.customer_id 
        
        response.append({
            "id": r.id,
            "full_name": r.full_name,
            "position": r.position or "",
            "gender": r.gender or "",
            "shift": r.shift or "",
            "room_number": room.room_number if room else "",
            "room_location": location.name if location else "",
            "room_path": room.path.description if room and room.path else "",
            "room_capacity": "",
            "field": r.field.name if r.field else "",
            "customer": r.customer.name if r.customer else "",
            "days_info": days_info
        })
        
    
    return response[:10]


@app.post("/api/add_resident")
async def add_resident(data: dict = Body(...), db: AsyncSession = Depends(get_db)):
    try:
        # 1. Месторождение
        res = await db.execute(select(Field).where(Field.name == data["field"]))
        field = res.scalars().first()
        if not field:
            field = Field(name=data["field"])
            db.add(field)

        # 2. Заказчик
        res = await db.execute(select(Customer).where(Customer.name == data["customer"]))
        customer = res.scalars().first()
        if not customer:
            customer = Customer(name=data["customer"])
            db.add(customer)

        # 3. Расположение
        res = await db.execute(select(Location).where(Location.name == data["location"]))
        location = res.scalars().first()
        if not location:
            location = Location(name=data["location"])
            db.add(location)

        # 4. Путь
        res = await db.execute(select(Path).where(Path.description == data["path"]))
        path = res.scalars().first()
        if not path:
            path = Path(description=data["path"])
            db.add(path)

        await db.flush()

        # 5. Комната – ИСПРАВЛЕНО: выбираем объект, а не только id
        raw_uid = data.get("room_unique_id", "")
        room_unique_id= raw_uid + 'a'
        capacity_val = int(raw_uid) if raw_uid.isdigit() else 0

        res = await db.execute(
            select(Room).where(
                and_(
                    Room.room_number == data['room_number'],
                    Room.room_unique_id == room_unique_id
                )
            )
        )
        room = res.scalars().first()
        if not room:
            room = Room(
                room_number=data['room_number'],
                field_id=field.id,
                capacity=capacity_val,
                location_id=location.id,
                path_id=path.id,
                room_unique_id=room_unique_id
            )
            db.add(room)
            await db.flush()

        # 6. Проживающий
        resident = Resident(
            field_id=field.id,
            customer_id=customer.id,
            full_name=data["full_name"],
            check_in=datetime.strptime(data["check_in"], "%Y-%m-%d").date(),
            check_out=datetime.strptime(data["check_out"], "%Y-%m-%d").date(),
            position=data.get("position", ""),
            gender=data.get("gender", ""),
            room_id=room.id,          # ← room – всегда объект
            shift=data.get("shift", "")
        )
        db.add(resident)
        await db.flush()

        # 7. День заезда (только одна запись, как ты и хотел)
        day = ResidentDay(
            resident_id=resident.id,
            room_id=resident.room_id,
            date=resident.check_in,
            extra=resident.check_out,
            customer_id=customer.id
        )
        db.add(day)
        await db.commit()

        return {"message": "Запись успешно добавлена", "resident_id": resident.id}

    except Exception as e:
        await db.rollback()
        # Возвращаем 400 с деталями ошибки
        raise HTTPException(status_code=400, detail=str(e))
    
@app.post("/api/upload_excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    print("=" * 60, flush=True)
    print(f"ЗАГРУЗКА ФАЙЛА: {file.filename}", flush=True)

    if not file.filename.endswith((".xlsx", ".xls")):
        return JSONResponse({"error": "Неверный формат файла"}, status_code=400)
    try:
        try:
            # Извлекаем имя месторождения и год из имени файла
            base_name = file.filename.rsplit(".", 1)[0].strip()
            parts = base_name.split("_")
            field_name = parts[0]
            year = int(parts[1]) if len(parts) > 1 else 2025
        except (IndexError, ValueError):
            return JSONResponse({"error": "Имя файла должно быть вида Урманское_2025.xlsx"}, status_code=400)

        try:
            contents = await file.read()
            wb = load_workbook(io.BytesIO(contents), data_only=False)

            # ШАГ 1: Читаем все записи из Excel в память
            all_records = list(parse_all_months(wb, year=year))
            print(f"  Распарсено из Excel: {len(all_records)} записей", flush=True)

            errors = []
            total = 0

            # ШАГ 2: Создаём все справочники (Field, Customer, Location, Path, Room)
            # Храним только id, не объекты SQLAlchemy
            customer_cache = {
                c.name: c.id
                for c in (await db.execute(select(Customer))).scalars()
            }

            location_cache = {
                l.name: l.id
                for l in (await db.execute(select(Location))).scalars()
            }

            path_cache = {
                p.description: p.id
                for p in (await db.execute(select(Path))).scalars()
            }

            workplace_cach = {
                w.name: w.id
                for w in (await db.execute(select(Workplace))).scalars()
            }

            room_cache = {
                w.room_number + w.room_unique_id: w.id
                for w in (await db.execute(select(Room))).scalars()
            }
            
            res = await db.execute(select(Field).where(Field.name == field_name))
            field = res.scalars().first()
            if not field:
                field = Field(name=field_name)
                db.add(field)
                await db.flush()
            field_id = field.id

        except Exception as e:
            await db.rollback()
            print(f"❌ Критическая ошибка: {e}", flush=True)
            return JSONResponse({"error": str(e)}, status_code=500)
        
        for rec in all_records:
            customer_name = rec.get("customer") or "Неизвестно"
            location_name = rec.get("расположение") or "-"
            path_desc     = rec.get("путь") or "-"
            room_name     = rec.get("комната") or "—"
            workplace_name     = rec.get("workplace") or "—"

            if workplace_name != "—" and workplace_name not in workplace_cach:
                workplace_obj = Workplace(name=workplace_name)
                db.add(workplace_obj)
                workplace_cach[workplace_name] = workplace_obj
            
            if customer_name not in customer_cache:
                customer_obj = Customer(name=customer_name)
                db.add(customer_obj)
                customer_cache[customer_name] = customer_obj

            if location_name not in location_cache:
                location_obj = Location(name=location_name)
                db.add(location_obj)
                location_cache[location_name] = location_obj

            if path_desc not in path_cache:
                path_obj = Path(description=path_desc)
                db.add(path_obj)
                path_cache[path_desc] = path_obj

        await db.flush()
        
        workplace_cach = {
            k: (v.id if isinstance(v, Workplace) else v)
            for k, v in workplace_cach.items()
        }
        customer_cache = {
            k: (v.id if isinstance(v, Customer) else v)
            for k, v in customer_cache.items()
        }
        location_cache = {
            k: (v.id if isinstance(v, Location) else v)
            for k, v in location_cache.items()
        }
        path_cache = {
            k: (v.id if isinstance(v, Path) else v)
            for k, v in path_cache.items()
        }
            
            
        try:
            for rec in all_records:
                room_name = rec.get("комната") or "—"
                room_uid = rec.get("room_unique_id") or ""
                room_cache_key = room_name + room_uid
                if room_cache_key in room_cache:
                    continue

                location_name = rec.get("расположение") or "-"
                path_desc = rec.get("путь") or "-"

                if room_cache_key not in room_cache:
                    room_obj = Room(
                        room_number=room_name,
                        capacity=int(rec.get("мест")) if rec.get("мест") else 0,
                        field_id=field_id,
                        location_id=location_cache[location_name],
                        path_id=path_cache[path_desc],
                        room_unique_id=room_uid or None
                    )
                    db.add(room_obj)
                    
                    room_cache[room_cache_key] = room_obj
        
        except Exception as e:
            await db.rollback()  
            print(f"Критическая ошибка: {e}", flush=True)
            return JSONResponse(
                {"error": str(e)},
                status_code=500
            )


        await db.flush()
        room_cache = {
                k: v.id if isinstance(v, Room) else v
                for k, v in room_cache.items()
            }
        await db.commit()
        print(f"  Справочники сохранены", flush=True)

        # ШАГ 3: Добавляем жильцов, каждый — отдельный commit
        # При ошибке одного делаем rollback только его, справочники уже в БД
        pending = 0
        for rec in all_records:
            try:
                workplace_id= workplace_cach.get(rec.get("workplace") or "—")
                customer_id = customer_cache[rec.get("customer") or "Неизвестно"]
                room_uid = rec.get("room_unique_id") or ""
                room_cache_key = (rec.get("комната") or "—") + room_uid
                room_id = room_cache[room_cache_key]
                gender_raw = rec.get("пол", "")
                gender = "М" if isinstance(gender_raw, str) and gender_raw.lower().startswith("муж") else "Ж"

                resident = Resident(
                    field_id=field_id,
                    customer_id=customer_id,
                    full_name=rec["full_name"],
                    position=rec.get("position", ""),
                    check_in=rec["check_in"],
                    check_out=rec["check_out"],
                    gender=gender,
                    room_id=room_id,
                    shift=rec.get("смена", ""),
                )
                db.add(resident)
                await db.flush()
                pending += 1
                
                
                db.add(ResidentDay(
                    resident_id=resident.id,
                    date=rec["check_in"],
                    extra=rec["check_out"],
                    customer_id=customer_id,
                    room_id=room_id,
                    workplace_id=workplace_id
                ))
                if pending >= 50:
                    await db.commit()
                    pending = 0
                
                total += 1
                # print(f"  ✓ [{total}] {rec['full_name']} {rec['check_in']} → {rec['check_out']}", flush=True)

            except Exception as e:
                await db.rollback()  
                errors.append({"record": rec.get("full_name", "?"), "error": str(e)})
                print(f"  ✗ Ошибка: {rec.get('full_name')}: {e}", flush=True)


        if pending:
            await db.commit()

        print(f"✅ Готово: {total} записей, ошибок: {len(errors)}", flush=True)
        return {
            "message": f"Загружено {total} записей" + (f", пропущено: {len(errors)}" if errors else ""),
            "errors": errors[:20],
        }
    except Exception as e:
        await db.rollback()  
        print(f"Критическая ошибка: {e}", flush=True)
        return JSONResponse(
            {"error": str(e)},
            status_code=500
        )


@app.get('/api/get_report')
async def get_report(date_in : date, date_out : date, db: AsyncSession = Depends(get_db)):
    
    # print(f"Ищем за период: {date_from} — {date_to}") 
    res = await db.execute(
                            select(Resident)
                           .join(Field, Field.id == Resident.field_id)
                           .join(Customer, Customer.id == Resident.customer_id)
                           .where(and_(Resident.check_in <= date_out,
                                       Resident.check_out >= date_in))
                            .order_by(Field.name, Customer.name)
                            .options(
                            selectinload(Resident.field),      # ← загружаем field
                            selectinload(Resident.customer)   # ← загружаем customer
                            )
                        )
    residents = res.scalars().unique().all()
    # print(f"Найдено жильцов: {len(residents)}")  

    if not residents:
        return JSONResponse({"error": "Нет данных за выбранный период"}, status_code=404)

    f = []
    for r in residents:
        actual_in = r.check_in if r.check_in >= date_in else date_in
        actual_out = r.check_out if r.check_out and r.check_out <= date_out else date_out
        days = (actual_out - actual_in).days + 1

        f.append({
            'Месторождение': r.field.name,
            'Заказчик': r.customer.name,
            'ФИО проживающего': r.full_name,
            'Дата заезда': actual_in,
            'Дата выезда': actual_out,
            'Количество дней': days
        })
    
    file_path = create_report(f, f"report_{date_in}_{date_out}.xlsx")
    return FileResponse(file_path, filename=os.path.basename(file_path))



@app.get("/api/fields")
async def get_fields(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Field))
    fields = result.scalars().all()
    return [
        {"id": f.id, "name": f.name}
        for f in fields
    ]


@app.get("/api/current_user")
def current_user(user: User = Depends(get_current_user)):
    return {"username": user.username, "role": user.role.name}

@app.get("/admin_management", response_class=HTMLResponse)
def admin_page(request: Request, user: User = Depends(admin_only)):
    return templates.TemplateResponse("admin_management.html", {"request": request})

@app.get("/api/get_admins")
async def get_admins(db: AsyncSession = Depends(get_db), user: User = Depends(admin_only)):
    res = await db.execute(select(User)
                           .join(Role, User.role_id == Role.id)
                           .where(Role.name == 'admin')
                           .options(selectinload(User.field)) )
    admins = res.scalars().unique().all()
    result = []
    for u in admins:
        # добавляем поле "field" если оно есть, иначе None
        field_name = u.field.name if hasattr(u, "field") and u.field else ""
        result.append({
            "id": u.id,
            "username": u.username,
            "field": field_name
        })
    return result

@app.post("/api/create_admin")
async def create_admin(request: Request, user: User = Depends(admin_only), db : AsyncSession = Depends(get_db)): 
    data = await request.json()
    username = data.get("username")
    password = data.get("password")
    field_id = data.get("field_id")
    if not username or not password:
        return JSONResponse({"error": "Все поля обязательны"}, status_code=400)


    res = await db.execute(select(User).where(User.username == username))
    if res.scalars().first():
        return JSONResponse({"error": "Логин уже существует"}, status_code=400)

    admin_role_id = await get_admin_role_id(db)
    hashed_password = get_password_hash(password)

    new_admin = User(
        username=username,
        password=hashed_password,
        role_id=admin_role_id,
        field_id=field_id if field_id else None,
    )
    db.add(new_admin)
    await db.commit()
    await db.refresh(new_admin)
    return {"message": f"Админ {username} создан!"}
    








@app.put("/api/update_admin_inline/{admin_id}")
async def update_admin_inline(admin_id: int, data: dict = Body(...), db : AsyncSession = Depends(get_db), current_user: User = Depends(admin_only)): ##############################
    try:
        res = await db.execute(select(User).where(User.id == admin_id))
        admin = res.scalars().first()
        if not admin:
            return JSONResponse({"error": "Админ не найден"}, status_code=404)

        if data.get("username"):
            admin.username = data["username"]
        if data.get("password"):
            admin.password = get_password_hash(data["password"])
        if data.get("field_id"):
            admin.field_id = data["field_id"]

        await db.commit()
        return {"message": "Обновлено"}
    except  Exception as e:
        await db.rollback()
        return JSONResponse(content={"status": "error", "detail": str(e)}, status_code=500)
    


@app.delete("/api/delete_admin/{admin_id}")
async def delete_admin(admin_id: int, db : AsyncSession = Depends(get_db), current_user: User = Depends(admin_only)):
    try:
        res = await db.execute(select(User).where(User.id == admin_id))
        admin = res.scalars().first()
        if not admin:
            return JSONResponse({"error": "Админ не найден"}, status_code=404)
        db.delete(admin)
        await db.commit()
        return {"message": "Админ удален"}
    except  Exception as e:
        await db.rollback()
        return JSONResponse(content={"status": "error", "detail": str(e)}, status_code=500)


@app.post("/api/update_resident")
async def update_resident(data: dict = Body(...), db: AsyncSession = Depends(get_db)):
    try:
        res = await db.execute(select(Resident).where(Resident.id == data["id"]))
        resident = res.scalars().first()
        if not resident:
            return JSONResponse(content={"error": "Жилец не найден"}, status_code=404)

        if "position" in data:
            resident.position = data["position"]
        if "gender" in data:
            resident.gender = data["gender"]
        if "shift" in data:
            resident.shift = data["shift"]

        await db.commit()
        await db.refresh(resident)
        return JSONResponse(content={"status": "ok"})
    except Exception as e:
        await db.rollback()
        return JSONResponse(content={"status": "error", "detail": str(e)}, status_code=500)


@app.get("/api/customers")
async def get_customers(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Customer))
    customers = res.scalars().all()
    return [{"id": c.id, "name": c.name} for c in customers]