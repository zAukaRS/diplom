from __future__ import annotations

import io
import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Generator, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("excel_parser")

CONTRACT_DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{2,4})")

# Excel (Windows) хранит даты как количество дней от 30.12.1899
# (с учётом фиктивного 29.02.1900 — отсюда именно эта база, а не 31.12.1899).
_EXCEL_EPOCH = date(1899, 12, 30)


def _clean_str(value) -> str:
    if value is None:
        return ""
    s = str(value).replace("\xa0", " ")
    s = "".join(" " if unicodedata.category(c) == "Zs" else c for c in s)
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _to_date(value) -> Optional[date]:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return _EXCEL_EPOCH + timedelta(days=int(value))

    value = _clean_str(value)
    if value in (None, ""):
        return None
    # ❗ защита от мусорных месяцев типа 05.13.2026
    if isinstance(value, str):
        if "." in value:
            parts = value.split(".")
            if len(parts) == 3:
                try:
                    _, m, _ = parts

                    if not (1 <= int(m) <= 12):
                        return None

                except Exception:
                    return None
    m = CONTRACT_DATE_RE.match(value)
    if not m:
        return None

    d, mo, y = map(int, m.groups())
    if y < 100:
        y += 2000
    
    try:
        return date(y, mo, d)
    except ValueError:
        return None



def _split_contract(raw):
    if not raw:
        return None, None

    raw = str(raw)

    m = CONTRACT_DATE_RE.search(raw)
    contract_date = None

    if m:
        d, mo, y = map(int, m.groups())
        if y < 100:
            y += 2000
        contract_date = date(y, mo, d)

    idx = raw.lower().find(" от ")
    contract_num = raw[:idx] if idx != -1 else raw
    return contract_num.strip(), contract_date



@dataclass
class ParsedRequestRow:
    full_name: str
    position: Optional[str]

    field_name: str

    check_in: date
    check_out: date
    days: Optional[int]

    customer_name: Optional[str]
    contract_num: Optional[str]
    contract_date: Optional[date]

    eol_fio: Optional[str]

    is_full_form: bool

    sheet_name: str
    row_number: int


def find_header_row(ws, max_scan_rows=20):
    max_row = min(ws.max_row, max_scan_rows)

    for row_idx in range(1, max_row):

        a = str(ws.cell(row_idx, 1).value or "").lower()
        d = str(ws.cell(row_idx, 4).value or "").lower()
        f = str(ws.cell(row_idx, 6).value or "").lower()

        g = ws.cell(row_idx + 1, 7).value
        h = ws.cell(row_idx + 1, 8).value

        score = 0

        if "наименование" in a:
            score += 1

        if "фио" in d or "работник" in d:
            score += 1

        if "объект" in f or "месторождение" in f:
            score += 1

        if _to_date(g):
            score += 1

        if _to_date(h):
            score += 1

        if score >= 4:
            logger.info("HEADER_FOUND: лист '%s': заголовок в строке %s (score=%s)", ws.title, row_idx, score)
            return row_idx

    logger.warning("HEADER_NOT_FOUND: лист '%s': шапка не найдена в первых %s строках", ws.title, max_scan_rows)
    return None


def parse_sheet(ws: Worksheet) -> Generator[ParsedRequestRow, None, None]:
    """
    Чтение листа.

    После определения шапки все данные читаются
    строго по индексам колонок:

        A организация
        B договор
        C ЕОЛ
        D ФИО
        E должность
        F месторождение
        G заезд
        H выезд
        I дни
    """

    header_row = find_header_row(ws)

    if header_row is None:
        raise ValueError(
            f"Лист '{ws.title}': не удалось определить строку заголовков "
            f"(в первых 20 строках не найдена структура заявки)"
        )

    header_found = True

    if not header_found:
        return

    parsed_count = 0
    skipped_empty = 0

    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row
    ):

        customer_name = row[0].value
        contract_raw = row[1].value
        eol_fio = row[2].value
        full_name = row[3].value
        position = row[4].value
        field_name = row[5].value

        check_in = _to_date(row[6].value)
        check_out = _to_date(row[7].value)

        days_raw = row[8].value

        # полностью пустая строка
        if all(
            cell.value in (None, "")
            for cell in row[:9]
        ):
            continue

        # служебные строки пропускаем
        if not full_name:
            skipped_empty += 1
            continue

        if not field_name:
            skipped_empty += 1
            continue

        if not check_in or not check_out:
            logger.error(
                "PARSE_ERROR invalid date: sheet=%s row=%s in=%r out=%r",
                ws.title, row[0].row, row[6].value, row[7].value
            )
            continue

        contract_num, contract_date = _split_contract(contract_raw)

        try:
            days = (
                int(days_raw)
                if days_raw not in (None, "")
                else (check_out - check_in).days
            )
        except Exception:
            days = (check_out - check_in).days

        cleaned_full_name = _clean_str(full_name)
        cleaned_field_name = _clean_str(field_name)

        if not cleaned_full_name or not cleaned_field_name:
            # после очистки строка оказалась пустой (например, ячейка
            # содержала только NBSP) — пропускаем как служебную
            continue
        if check_in > check_out:
            logger.error(
                "PARSE_ERROR inverted dates: %s row=%s %s > %s",
                ws.title, row[0].row, check_in, check_out
            )
            continue
        actual_days = (check_out - check_in).days

        if days != actual_days:
            logger.warning(
                "DAYS_MISMATCH: %s row=%s excel=%s actual=%s",
                ws.title,
                row[0].row,
                days,
                actual_days,
            )
        
        yield ParsedRequestRow(
            full_name=cleaned_full_name,

            position=(
                _clean_str(position)
                if position
                else None
            ) or None,

            field_name=cleaned_field_name,

            check_in=check_in,
            check_out=check_out,

            days=days,

            customer_name=(
                _clean_str(customer_name)
                if customer_name
                else None
            ) or None,

            contract_num=contract_num,
            contract_date=contract_date,

            eol_fio=(
                _clean_str(eol_fio)
                if eol_fio
                else None
            ) or None,

            is_full_form = bool(contract_raw and str(contract_raw).strip()),

            sheet_name=ws.title,

            row_number=row[0].row,
        )

        parsed_count += 1

    logger.info(
        "PARSE_SHEET: лист '%s': распознано строк=%s, пропущено пустых/служебных=%s",
        ws.title, parsed_count, skipped_empty,
    )


def parse_workbook(wb_or_path):
    """
    Перебирает все листы книги.
    """

    if isinstance(wb_or_path, (str, bytes)):
        wb = load_workbook(
            wb_or_path,
            data_only=True
        )

    elif hasattr(wb_or_path, "read"):
        wb = load_workbook(
            io.BytesIO(wb_or_path.read()),

            
            data_only=True
        )

    else:
        wb = wb_or_path

    logger.info("PARSE_WORKBOOK: листов в книге=%s: %s", len(wb.sheetnames), wb.sheetnames)

    total = 0
    for sheet_name in wb.sheetnames:
        for parsed_row in parse_sheet(wb[sheet_name]):
            total += 1
            yield parsed_row

    logger.info("PARSE_WORKBOOK: всего строк распознано по всем листам=%s", total)