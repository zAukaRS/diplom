"""
Упрощённый парсер Excel для импорта Location / Path / Room.

Поддерживает два формата листа:

  Формат A (основной парсер — колонки A/B/C/D):
    A – Расположение  (ffill)
    B – Путь          (ffill)
    C – № комнаты     (ffill)
    D – К-во мест     (ffill)

  Формат B (файл Urman — колонки A/B/C без Пути):
    A – Расположение  (ffill)
    B – № комнаты     (ffill)
    C – К-во мест     (ffill)

Формат определяется автоматически по заголовку первой строки.
Если колонка B содержит числа (не текст) — это Формат B.

Добавлено: остановка парсинга при 15+ пустых строках подряд.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Типы данных результата
# ---------------------------------------------------------------------------

@dataclass
class LocationRow:
    name: str

@dataclass
class PathRow:
    description: str

@dataclass
class RoomRow:
    room_number: str
    capacity: int
    room_unique_id: str
    location_name: str
    path_description: str
    gender: Optional[str] = None    # если есть колонка Пол
    status: int = 1                 # 1 = активна


@dataclass
class ImportResult:
    locations: list[LocationRow] = field(default_factory=list)
    paths: list[PathRow] = field(default_factory=list)
    rooms: list[RoomRow] = field(default_factory=list)
    skipped_rows: int = 0
    errors: list[str] = field(default_factory=list)

    def summary(self) -> str:
        return (
            f"Locations: {len(self.locations)}, "
            f"Paths: {len(self.paths)}, "
            f"Rooms: {len(self.rooms)}, "
            f"Skipped: {self.skipped_rows}, "
            f"Errors: {len(self.errors)}"
        )


# ---------------------------------------------------------------------------
# Определение формата листа
# ---------------------------------------------------------------------------

def _detect_format(header_row: tuple) -> str:
    """
    Возвращает 'A' (есть колонка Путь) или 'B' (нет колонки Путь).

    Формат A: заголовки вроде ['Расположение', 'Путь', '№ комнаты', ...]
    Формат B: заголовки вроде ['Расположение', '№ комнаты', 'К-во мест', ...]
    """
    if not header_row:
        return 'B'

    # Нормализуем заголовки
    headers = [str(h).strip().lower() if h else "" for h in header_row[:5]]

    path_keywords = ("путь", "path", "маршрут", "дорога")
    for h in headers:
        if any(kw in h for kw in path_keywords):
            return 'A'

    return 'B'


# ---------------------------------------------------------------------------
# Основной парсер
# ---------------------------------------------------------------------------

DEFAULT_LOCATION = "Общежитие"   # если путь есть, а расположение — нет
DEFAULT_PATH     = "Без пути"    # если расположение есть, а путь — нет

# Максимальное количество пустых строк подряд, после которых парсинг прекращается
MAX_EMPTY_ROWS = 15


def import_rooms_from_excel(
    path_or_file,
    field_id: int,
    sheet_name: str | None = None,
    default_path: str = "Основной",
    dry_run: bool = True,
) -> ImportResult:
    """
    Парсит Excel и возвращает ImportResult со списками объектов для создания.

    Args:
        path_or_file:   Путь к файлу, bytes или file-like объект.
        field_id:       ID месторождения (Field) в БД — только для справки,
                        не используется при dry_run=True.
        sheet_name:     Имя листа. None = первый лист.
        default_path:   Путь по умолчанию для формата B (где нет колонки Path).
        dry_run:        Если True — только парсить, не сохранять.

    Returns:
        ImportResult
    """
    # Загрузка книги
    if isinstance(path_or_file, (str, bytes)):
        wb = load_workbook(path_or_file, data_only=True)
    elif hasattr(path_or_file, "read"):
        content = path_or_file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    else:
        wb = path_or_file

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    result = ImportResult()

    # Множества для дедупликации
    seen_locations: set[str] = set()
    seen_paths: set[str] = set()
    # room_unique_id -> RoomRow для дедупликации комнат
    seen_rooms: dict[str, RoomRow] = {}

    # --- Читаем заголовок ---
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        result.errors.append("Лист пустой")
        return result

    fmt = _detect_format(header)

    # Маппинг колонок (0-based)
    if fmt == 'A':
        COL_LOCATION = 0
        COL_PATH     = 1
        COL_ROOM     = 2
        COL_SEATS    = 3
        COL_GENDER   = 4   # необязательно
    else:  # fmt == 'B'
        COL_LOCATION = 0
        COL_PATH     = None   # нет
        COL_ROOM     = 1
        COL_SEATS    = 2
        COL_GENDER   = 3   # необязательно

    # Контекстные переменные (ffill)
    ctx_location = ""        # "" = ещё не встретили расположение
    ctx_path     = ""        # "" = ещё не встретили путь
    ctx_room     = ""
    ctx_seats    = 1
    ctx_gender   = None
    list_of_rooms: list[str] = []   # для room_unique_id внутри одной комнаты

    def cv(row, idx) -> str | None:
        """Вернуть значение ячейки как строку или None."""
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    def cv_int(row, idx) -> int | None:
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        try:
            return int(v)
        except (ValueError, TypeError):
            return None

    # Счётчик последовательных пустых строк
    consecutive_empty = 0

    for row_num, row in enumerate(rows_iter, start=2):
        # --- Проверка, является ли строка полностью пустой (по данным ячеек) ---
        # Извлекаем значения для проверки
        loc_val = cv(row, COL_LOCATION)
        path_val = cv(row, COL_PATH) if fmt == 'A' else None
        room_val = cv(row, COL_ROOM)
        seats_val = cv_int(row, COL_SEATS)
        gender_val = cv(row, COL_GENDER)

        # Строка считается пустой, если все значимые поля отсутствуют
        is_empty = (
            not loc_val and
            not path_val and
            not room_val and
            seats_val is None and
            not gender_val
        )

        if is_empty:
            consecutive_empty += 1
            if consecutive_empty > MAX_EMPTY_ROWS:
                # Слишком много пустых строк подряд — прекращаем парсинг
                result.errors.append(
                    f"Превышено количество пустых строк ({MAX_EMPTY_ROWS}) подряд. "
                    f"Парсинг остановлен на строке {row_num}."
                )
                break
            # Пропускаем пустую строку
            result.skipped_rows += 1
            continue
        else:
            # Сброс счётчика при нахождении непустой строки
            consecutive_empty = 0

        # --- Обработка непустой строки ---
        # ffill
        if loc_val:
            # Убираем лишние переносы строк (как '\nАБЖК')
            loc_val = loc_val.replace("\n", "").strip()
            if loc_val.upper() == "РАСПОЛОЖЕНИЕ":
                # Повторный заголовок — пропускаем
                result.skipped_rows += 1
                continue
            ctx_location = loc_val
            list_of_rooms = []   # новое расположение — сбрасываем счётчик комнат

        if fmt == 'A' and path_val:
            if path_val != ctx_path:
                # Новый путь — если расположение в этой строке не задано,
                # сбрасываем ctx_location чтобы сработал фолбэк "Общежитие"
                if not loc_val:   # если в этой строке нет расположения
                    ctx_location = ""
                ctx_path = path_val
                list_of_rooms = []

        if room_val:
            ctx_room = room_val
            list_of_rooms = []

        if seats_val is not None:
            ctx_seats = seats_val

        if gender_val:
            ctx_gender = gender_val

        # --- Спецпроверка: К-во мест > 10, и при этом в колонке с индексом 18
        # (0-based, т.е. 19-я колонка листа) стоит число — считаем такую строку
        # служебной/ошибочной (не реальная комната) и пропускаем её ---
        extra_col_val = cv_int(row, 18)
        if ctx_seats is not None and ctx_seats > 10 and extra_col_val is not None:
            result.skipped_rows += 1
            continue

        # --- Пропускаем строки без места ---
        if seats_val is None and not room_val:
            result.skipped_rows += 1
            continue

        # Нет ни расположения ни пути ни комнаты — пустая строка (уже отловлена выше)
        if not ctx_room and not ctx_location and not ctx_path:
            result.skipped_rows += 1
            continue

        if not ctx_room:
            result.skipped_rows += 1
            continue

        # --- Эффективные значения с фолбэками ---
        # Путь есть, расположения нет → location = "Общежитие"
        # Расположение есть, пути нет  → path = "Без пути"
        eff_location = ctx_location if ctx_location else DEFAULT_LOCATION
        eff_path     = ctx_path     if ctx_path     else DEFAULT_PATH

        # --- Генерируем room_unique_id ---
        base_uid = str(ctx_room)
        if room_val:
            # Это первое место в этой комнате
            uid = base_uid + 'a'
            list_of_rooms.append(uid)
        else:
            # Дополнительное место в той же комнате
            uid = None
            for suffix in 'bcdefghij':
                candidate = base_uid + suffix
                if candidate not in list_of_rooms:
                    uid = candidate
                    list_of_rooms.append(uid)
                    break
            if uid is None:
                result.errors.append(
                    f"Строка {row_num}: не смогли назначить room_unique_id для комнаты {ctx_room}"
                )
                continue

        # --- Дедупликация Location ---
        if eff_location not in seen_locations:
            seen_locations.add(eff_location)
            result.locations.append(LocationRow(name=eff_location))

        # --- Дедупликация Path ---
        if eff_path not in seen_paths:
            seen_paths.add(eff_path)
            result.paths.append(PathRow(description=eff_path))

        # --- Дедупликация Room ---
        if uid not in seen_rooms:
            room_row = RoomRow(
                room_number=ctx_room,
                capacity=ctx_seats,
                room_unique_id=uid,
                location_name=eff_location,
                path_description=eff_path,
                gender=ctx_gender,
                status=1,
            )
            seen_rooms[uid] = room_row
            result.rooms.append(room_row)

    return result


# ---------------------------------------------------------------------------
# Сохранение в БД (синхронная версия — для совместимости)
# ---------------------------------------------------------------------------

def apply_to_db(
    result: ImportResult,
    db,
    field_id: int,
    skip_existing: bool = True,
) -> dict:
    """
    Создаёт объекты Location / Path / Room в БД.

    Args:
        result:         Результат import_rooms_from_excel().
        db:             SQLAlchemy Session.
        field_id:       ID месторождения.
        skip_existing:  Если True — пропускает уже существующие по имени/описанию.

    Returns:
        Словарь со статистикой: {'locations': N, 'paths': N, 'rooms': N, 'skipped': N}
    """
    from app.models import Location, Path, Room

    stats = {"locations": 0, "paths": 0, "rooms": 0, "skipped": 0}

    # --- Locations ---
    location_map: dict[str, int] = {}
    for loc_row in result.locations:
        existing = db.query(Location).filter(Location.name == loc_row.name).first()
        if existing:
            location_map[loc_row.name] = existing.id
            if skip_existing:
                stats["skipped"] += 1
                continue
        loc = Location(name=loc_row.name)
        db.add(loc)
        db.flush()
        location_map[loc_row.name] = loc.id
        stats["locations"] += 1

    # --- Paths ---
    path_map: dict[str, int] = {}
    for path_row in result.paths:
        existing = db.query(Path).filter(Path.description == path_row.description).first()
        if existing:
            path_map[path_row.description] = existing.id
            if skip_existing:
                stats["skipped"] += 1
                continue
        p = Path(description=path_row.description)
        db.add(p)
        db.flush()
        path_map[path_row.description] = p.id
        stats["paths"] += 1

    # --- Rooms ---
    for room_row in result.rooms:
        existing = db.query(Room).filter(
            Room.room_unique_id == room_row.room_unique_id,
            Room.field_id == field_id,
        ).first()
        if existing and skip_existing:
            stats["skipped"] += 1
            continue

        loc_id = location_map.get(room_row.location_name)
        path_id = path_map.get(room_row.path_description)

        r = Room(
            room_number=room_row.room_number,
            capacity=room_row.capacity,
            room_unique_id=room_row.room_unique_id,
            field_id=field_id,
            location_id=loc_id,
            path_id=path_id,
            status=room_row.status,
        )
        db.add(r)
        stats["rooms"] += 1

    db.commit()
    return stats


# ---------------------------------------------------------------------------
# CLI / быстрая проверка
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else None
    sheet = sys.argv[2] if len(sys.argv) > 2 else None

    if not path:
        print("Использование: python simple_room_importer.py <файл.xlsx> [лист]")
        sys.exit(1)

    result = import_rooms_from_excel(path, field_id=0, sheet_name=sheet, dry_run=True)
    print(result.summary())
    print()
    print("=== Locations ===")
    for loc in result.locations:
        print(f"  {loc.name}")
    print()
    print("=== Paths ===")
    for p in result.paths:
        print(f"  {p.description}")
    print()
    print("=== Rooms (первые 20) ===")
    for r in result.rooms[:20]:
        print(f"  [{r.room_unique_id}] комната {r.room_number}, "
              f"{r.capacity} мест, loc={r.location_name}, path={r.path_description}")
    if len(result.rooms) > 20:
        print(f"  ... и ещё {len(result.rooms) - 20}")
    if result.errors:
        print()
        print("=== Ошибки ===")
        for e in result.errors:
            print(f"  ⚠ {e}")